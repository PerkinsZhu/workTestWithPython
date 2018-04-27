import xlrd
import xlwt
import unittest
import os
import time
from openpyxl import load_workbook
from openpyxl import Workbook


class Excel(unittest.TestCase):

    def test_Read(self):
        data = Excel.loadData()
        sheet: xlrd.sheet.Sheet
        for sheet in data.sheets():
            print(sheet.nrows)
            print(sheet.ncols)
            for row in range(0, sheet.nrows):
                print(sheet.row_values(row))

    @staticmethod
    def loadData():
        # return xlrd.open_workbook("G:\\test\\chognqihoutai1.xlsx")
        return xlrd.open_workbook("G:\\test\\python.xlsx")

    def testList(self):
        for i in [2, "34", 345]:
            print(i)

    def testExcel(self):
        file = "G:\\test\\python.xls"
        # excelFile = unicode(file, "UTF-8")
        if (os._exists(file)):
            os.remove(file)
        wbk = xlwt.Workbook()
        sheet = wbk.add_sheet("sheet1", cell_overwrite_ok=True)
        title = ['name', 'age', 'phone']
        for i, v in enumerate(title):
            sheet.write(0, i, v)
        wbk.save(file)
        time.sleep(2)

    def testSaveXLSX(self):
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 3500
        sheet.append(["我", "你", "她"])
        sheet.append(["我", "你", "她"])
        sheet.append(["我", "你", "她"])
        wb.save("G:\\test\\python.xlsx")
